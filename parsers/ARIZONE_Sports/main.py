import os
import json
import re
import time
from io import BytesIO

import requests
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import sys


# ============================================================
# НАСТРОЙКИ
# ============================================================

BASE = "https://arizonesports.com.ua/uk"
DOMAIN = "arizonesports.com.ua"

GOOGLE_SHEET_ID = "1ROCGu-3W8PotpQSIgzHRDkXCiDeHU0_dzEjuRnkNHv0"

# Для теста
CATEGORY_LIMIT = 2
PRODUCT_LIMIT = 20

OUTPUT_DIR = os.path.abspath("output/ARIZONE_Sports")
FILE_PATH = os.path.join(OUTPUT_DIR, "ARIZONE_Sports_LIVE.xlsx")
STATUS_PATH = os.path.join(OUTPUT_DIR, "status.json")
LOCK_FILE = os.path.join(OUTPUT_DIR, "parser.lock")

USER = sys.argv[1] if len(sys.argv) > 1 else "unknown"


# ============================================================
# SESSION
# ============================================================

session = requests.Session()

session.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/140.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "uk-UA,uk;q=0.9,ru;q=0.8,en;q=0.7",
})


# ============================================================
# ПАПКИ
# ============================================================

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ============================================================
# LOCK
# ============================================================

def create_lock():
    if os.path.exists(LOCK_FILE):
        try:
            age = time.time() - os.path.getmtime(LOCK_FILE)

            if age < 3600:
                print("⚠️ Парсер уже запущен.")
                return False

            print("⚠️ Найден старый lock-файл. Удаляем.")
            os.remove(LOCK_FILE)

        except Exception:
            pass

    try:
        with open(LOCK_FILE, "w", encoding="utf-8") as f:
            f.write(str(os.getpid()))
        return True

    except Exception as e:
        print(f"❌ Не удалось создать lock: {e}")
        return False


def remove_lock():
    try:
        if os.path.exists(LOCK_FILE):
            os.remove(LOCK_FILE)
    except Exception:
        pass


# ============================================================
# STATUS
# ============================================================

def save_status(
    status,
    current=0,
    total=0,
    message="",
):
    data = {
        "status": status,
        "current": current,
        "total": total,
        "message": message,
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    try:
        with open(STATUS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️ STATUS ERROR: {e}")


# ============================================================
# HTTP
# ============================================================

def get_soup(url, timeout=40):
    try:
        response = session.get(
            url,
            timeout=timeout,
            allow_redirects=True,
        )

        response.raise_for_status()

        return BeautifulSoup(response.text, "html.parser")

    except Exception as e:
        print(f"❌ GET ERROR: {url}")
        print(f"   {e}")
        return None


# ============================================================
# HELPERS
# ============================================================

def clean(value):
    if value is None:
        return ""

    value = str(value)

    value = value.replace("\xa0", " ")
    value = value.replace("\u200b", "")

    return re.sub(r"\s+", " ", value).strip()


def absolute_url(url):
    if not url:
        return ""

    url = str(url).strip()

    if url.startswith("http://") or url.startswith("https://"):
        return url

    if url.startswith("//"):
        return "https:" + url

    if url.startswith("/"):
        return "https://" + DOMAIN + url

    return BASE.rstrip("/") + "/" + url.lstrip("/")


def base_product_url(url):
    """
    Убираем hash (#/цвет-...) и query,
    чтобы разные цвета считались одним товаром.
    """
    if not url:
        return ""

    url = url.split("#", 1)[0]
    url = url.split("?", 1)[0]

    return url.rstrip("/")


# ============================================================
# TRANSLITERATION
# ============================================================

TRANSLIT = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "ґ": "g",
    "д": "d",
    "е": "e",
    "є": "ye",
    "ж": "zh",
    "з": "z",
    "и": "y",
    "і": "i",
    "ї": "yi",
    "й": "y",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "kh",
    "ц": "ts",
    "ч": "ch",
    "ш": "sh",
    "щ": "shch",
    "ь": "",
    "ю": "yu",
    "я": "ya",
    "ы": "y",
    "э": "e",
    "ё": "yo",
    "ъ": "",
}


def translit_slug(text):
    text = clean(text).lower()

    result = []

    for char in text:
        result.append(TRANSLIT.get(char, char))

    text = "".join(result)

    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text)
    text = text.strip("-")

    return text


# ============================================================
# GOOGLE SHEETS
# ============================================================

def normalize_header(value):
    if value is None:
        return ""

    value = str(value)

    value = value.replace("\xa0", " ")
    value = value.replace("\u200b", "")

    value = re.sub(r"\s+", " ", value)

    return value.strip().casefold()


def normalize_sku(value):
    """
    Нормализация артикула.

    Например:
    NN988
    nn988
    NN988
    NN 988

    -> nn988
    """

    if value is None:
        return ""

    value = str(value)

    value = value.replace("\xa0", "")
    value = value.replace("\u200b", "")
    value = value.strip()

    # Excel иногда отдаёт числовые артикулы как 1918.0
    if re.fullmatch(r"\d+\.0+", value):
        value = value.split(".", 1)[0]

    # Убираем пробелы внутри артикула
    value = re.sub(r"\s+", "", value)

    return value.casefold()


def parse_price(value):
    """
    Преобразует значение цены в float.

    Поддерживает:
    18
    18.5
    18,5
    1 250
    1 250,50
    1.250,50
    """

    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    value = str(value).strip()

    if not value:
        return None

    value = value.replace("\xa0", " ")
    value = value.replace("$", "")
    value = value.replace("USD", "")
    value = value.strip()

    # Убираем обычные пробелы-разделители тысяч
    value = value.replace(" ", "")

    if "," in value and "." in value:

        # 1.250,50
        if value.rfind(",") > value.rfind("."):
            value = value.replace(".", "")
            value = value.replace(",", ".")

        # 1,250.50
        else:
            value = value.replace(",", "")

    elif "," in value:
        value = value.replace(",", ".")

    try:
        return float(value)

    except Exception:
        return None


def find_header_columns(header_row):
    """
    Ищем ТОЛЬКО:

    Артикул
    Оптова ціна, $

    Остальные колонки игнорируем.
    """

    article_col = None
    price_col = None

    for index, value in enumerate(header_row):

        header = normalize_header(value)

        if header == normalize_header("Артикул"):
            article_col = index

        elif header == normalize_header("Оптова ціна, $"):
            price_col = index

    return article_col, price_col


def get_google_sheet_price_map():
    """
    Загружает ВСЮ Google Таблицу одним XLSX-файлом.

    Затем перебирает ВСЕ вкладки.

    На каждой вкладке ищет:
        Артикул
        Оптова ціна, $

    Возвращает:
        {
            "nn988": 18.0,
            "bbtt-q7": 28.0,
            ...
        }
    """

    print("")
    print("=" * 70)
    print("📊 ЗАГРУЗКА ЦЕН ИЗ GOOGLE SHEETS")
    print("=" * 70)

    export_url = (
        "https://docs.google.com/spreadsheets/d/"
        f"{GOOGLE_SHEET_ID}"
        "/export?format=xlsx"
    )

    print("🌐 Скачиваем всю Google Таблицу...")
    print(f"   {export_url}")

    try:
        response = session.get(
            export_url,
            timeout=90,
            allow_redirects=True,
        )

        print(f"📥 HTTP: {response.status_code}")
        print(f"📦 Размер файла: {len(response.content):,} bytes")

        response.raise_for_status()

    except Exception as e:
        print("❌ Не удалось скачать Google Таблицу.")
        print(f"   {e}")
        return {}

    # Проверяем, действительно ли пришёл XLSX
    if not response.content.startswith(b"PK"):
        print("")
        print("❌ Google не вернула XLSX-файл.")
        print("⚠️ Вместо него получен другой ответ.")
        print("")
        print("Первые 300 символов ответа:")

        try:
            print(response.text[:300])
        except Exception:
            print("<не удалось прочитать ответ как текст>")

        return {}

    try:
        workbook = load_workbook(
            BytesIO(response.content),
            data_only=True,
            read_only=True,
        )

    except Exception as e:
        print("❌ Ошибка открытия XLSX Google Таблицы:")
        print(f"   {e}")
        return {}

    print("")
    print(f"📚 Найдено вкладок: {len(workbook.sheetnames)}")

    prices = {}

    duplicate_count = 0
    sheets_with_prices = 0
    total_price_rows = 0

    # ========================================================
    # ВСЕ ВКЛАДКИ
    # ========================================================

    for sheet_name in workbook.sheetnames:

        print("")
        print(f"📄 Лист: {sheet_name}")

        try:
            ws = workbook[sheet_name]

            row_iterator = ws.iter_rows(values_only=True)

            header_row = None
            article_col = None
            price_col = None

            # Ищем заголовки в первых 30 строках
            for row_number in range(1, 31):

                try:
                    row = next(row_iterator)
                except StopIteration:
                    break

                found_article, found_price = find_header_columns(row)

                if found_article is not None and found_price is not None:

                    header_row = row_number
                    article_col = found_article
                    price_col = found_price

                    break

            if header_row is None:

                print(
                    "   ⚠️ Не найдены колонки "
                    "'Артикул' и 'Оптова ціна, $'"
                )

                continue

            print(f"   ✅ Строка заголовков: {header_row}")
            print(f"   🔹 Артикул: колонка {article_col + 1}")
            print(f"   🔹 Оптова ціна, $: колонка {price_col + 1}")

            sheet_price_count = 0

            # ------------------------------------------------
            # Данные после строки заголовков
            # ------------------------------------------------

            for row in row_iterator:

                if not row:
                    continue

                if article_col >= len(row):
                    continue

                if price_col >= len(row):
                    continue

                sku_raw = row[article_col]
                price_raw = row[price_col]

                sku = normalize_sku(sku_raw)

                if not sku:
                    continue

                price = parse_price(price_raw)

                if price is None:
                    continue

                total_price_rows += 1
                sheet_price_count += 1

                if sku in prices:

                    duplicate_count += 1

                    print(
                        f"   ⚠️ Дубликат SKU: {sku_raw} "
                        f"→ уже {prices[sku]}, "
                        f"новая цена {price}"
                    )

                    # Оставляем первую найденную цену
                    continue

                prices[sku] = price

            if sheet_price_count > 0:
                sheets_with_prices += 1

            print(
                f"   💰 Цен найдено на листе: "
                f"{sheet_price_count}"
            )

        except Exception as e:

            print(
                f"   ❌ Ошибка обработки листа "
                f"{sheet_name}: {e}"
            )

    # ========================================================
    # РЕЗУЛЬТАТ
    # ========================================================

    print("")
    print("-" * 70)
    print(f"📚 Обработано вкладок: {len(workbook.sheetnames)}")
    print(f"📄 Вкладок с ценами: {sheets_with_prices}")
    print(f"💵 Строк с ценами: {total_price_rows}")
    print(f"🔁 Дубликатов SKU: {duplicate_count}")
    print(f"💰 Уникальных SKU с ценой: {len(prices)}")
    print("-" * 70)

    if prices:

        print("")
        print("🔎 Примеры загруженных цен:")

        for index, (sku, price) in enumerate(prices.items()):

            print(f"   {sku} → ${price:g}")

            if index >= 19:
                break

    else:

        print("")
        print("❌ НИ ОДНОЙ ЦЕНЫ НЕ ЗАГРУЖЕНО!")
        print("")
        print(
            "Проверь доступ Google Таблицы "
            "и наличие колонок:"
        )
        print("   Артикул")
        print("   Оптова ціна, $")

    print("")

    return prices


# ============================================================
# CATEGORIES
# ============================================================

def get_categories():
    """
    Получаем ВСЕ категории и подкатегории
    из дерева categories_block_left.

    Глубина не ограничивается.
    """

    print("")
    print("=" * 70)
    print("📂 ПОИСК КАТЕГОРИЙ")
    print("=" * 70)

    soup = get_soup(BASE)

    if not soup:
        return []

    categories = []

    tree = soup.select(
        "#categories_block_left ul.tree a[href]"
    )

    seen = set()

    for a in tree:

        href = a.get("href")

        if not href:
            continue

        url = absolute_url(href)
        url = url.split("#", 1)[0]

        # Не берём повторно
        if url in seen:
            continue

        seen.add(url)

        name = clean(a.get_text(" ", strip=True))

        if not name:
            continue

        categories.append({
            "name": name,
            "url": url,
        })

    print(f"📂 Найдено категорий/подкатегорий: {len(categories)}")

    for index, category in enumerate(categories, start=1):

        print(
            f"   {index}. "
            f"{category['name']} → {category['url']}"
        )

    return categories


# ============================================================
# SHOW ALL
# ============================================================

def get_show_all_data(soup):
    """
    На некоторых категориях есть форма:

    n=85
    id_category=14

    Используем её, чтобы попытаться получить
    все товары одной страницей.
    """

    if not soup:
        return None

    form = soup.select_one("form.showall")

    if not form:
        return None

    n_input = form.select_one("input[name='n']")
    id_input = form.select_one("input[name='id_category']")

    if not n_input or not id_input:
        return None

    n = clean(n_input.get("value"))
    category_id = clean(id_input.get("value"))

    if not n or not category_id:
        return None

    return {
        "n": n,
        "id_category": category_id,
    }


# ============================================================
# PRODUCTS
# ============================================================

def get_product_links(soup):

    if not soup:
        return []

    selectors = [
        "ul.product_list li.ajax_block_product a.product-name",
        ".product-container a.product-name",
        "li.product-box.item a.product-image",
        "li.product-box.item h5.product-name a",
        "li.product-box.item a[href*='.html']",
        "#productscategory_list li.product-box a[href*='.html']",
    ]

    links = []
    seen = set()

    for selector in selectors:

        for a in soup.select(selector):

            href = a.get("href")

            if not href:
                continue

            url = absolute_url(href)
            url = base_product_url(url)

            if not url:
                continue

            if url in seen:
                continue

            seen.add(url)
            links.append(url)

    return links


# ============================================================
# PAGINATION
# ============================================================

def get_pagination_pages(soup):

    if not soup:
        return []

    pages = []

    for a in soup.select(
        "#pagination_bottom ul.pagination a[href]"
    ):

        href = a.get("href")

        if not href:
            continue

        text = clean(a.get_text(" ", strip=True))

        match = re.search(
            r"#/page-(\d+)",
            href,
            re.I,
        )

        if not match:
            continue

        page_number = int(match.group(1))

        pages.append(page_number)

    return sorted(set(pages))


# ============================================================
# CATEGORY PRODUCTS
# ============================================================

def get_category_products(category_url):

    print("")
    print(f"📂 Категория: {category_url}")

    all_products = []
    seen_products = set()

    first_soup = get_soup(category_url)

    if not first_soup:
        return []

    # ========================================================
    # Сначала пробуем SHOW ALL
    # ========================================================

    show_all = get_show_all_data(first_soup)

    if show_all:

        try:

            show_all_url = (
                category_url
                + "?n="
                + show_all["n"]
                + "&id_category="
                + show_all["id_category"]
            )

            print(
                f"📄 Пробуем SHOW ALL: "
                f"{show_all_url}"
            )

            soup = get_soup(show_all_url)

            if soup:

                products = get_product_links(soup)

                for product_url in products:

                    if product_url in seen_products:
                        continue

                    seen_products.add(product_url)
                    all_products.append(product_url)

                print(
                    f"   🛒 SHOW ALL товаров: "
                    f"{len(products)}"
                )

                if all_products:
                    return all_products

        except Exception as e:

            print(f"⚠️ SHOW ALL ERROR: {e}")

    # ========================================================
    # Обычная первая страница
    # ========================================================

    products = get_product_links(first_soup)

    for product_url in products:

        if product_url in seen_products:
            continue

        seen_products.add(product_url)
        all_products.append(product_url)

    print(
        f"   🛒 Товаров на первой странице: "
        f"{len(products)}"
    )

    # ========================================================
    # Pagination
    # ========================================================

    pages = get_pagination_pages(first_soup)

    if pages:

        print(
            f"   📄 Найдены страницы: "
            f"{', '.join(map(str, pages))}"
        )

    for page in pages:

        if page <= 1:
            continue

        # ----------------------------------------------------
        # Вариант 1 — hash URL преобразуем в ?page=N
        # ----------------------------------------------------

        page_url = category_url

        if "?" in page_url:
            page_url += f"&page={page}"
        else:
            page_url += f"?page={page}"

        soup = get_soup(page_url)

        if not soup:
            continue

        page_products = get_product_links(soup)

        print(
            f"   📄 Страница {page}: "
            f"{len(page_products)} товаров"
        )

        for product_url in page_products:

            if product_url in seen_products:
                continue

            seen_products.add(product_url)
            all_products.append(product_url)

    print(
        f"🛒 Всего товаров категории: "
        f"{len(all_products)}"
    )

    return all_products


# ============================================================
# PRODUCT TITLE
# ============================================================

def get_product_title(soup):

    if not soup:
        return ""

    selectors = [
        "h1[itemprop='name']",
        "h1.product-name",
        "#product_name",
        "h1",
    ]

    for selector in selectors:

        element = soup.select_one(selector)

        if element:

            title = clean(
                element.get_text(
                    " ",
                    strip=True,
                )
            )

            if title:
                return title

    return ""


# ============================================================
# SKU
# ============================================================

def get_product_sku(soup):

    if not soup:
        return ""

    selectors = [
        "#product_reference span[itemprop='sku']",
        "#product_reference .editable",
        "#product_reference",
        "[itemprop='sku']",
    ]

    for selector in selectors:

        element = soup.select_one(selector)

        if element:

            sku = clean(
                element.get_text(
                    " ",
                    strip=True,
                )
            )

            if sku:

                # Иногда текст содержит "Артикул: NN988"
                sku = re.sub(
                    r"^(артикул|код|sku)\s*[:#-]?\s*",
                    "",
                    sku,
                    flags=re.I,
                )

                return clean(sku)

    return ""


# ============================================================
# STATUS
# ============================================================

def get_product_status(soup):

    if not soup:
        return "Немає в наявності"

    # --------------------------------------------------------
    # Видимый availability
    # --------------------------------------------------------

    element = soup.select_one(
        "#availability_value"
    )

    if element:

        text = clean(
            element.get_text(
                " ",
                strip=True,
            )
        ).casefold()

        unavailable_words = [
            "відсут",
            "немає",
            "нема",
            "немає в наявності",
            "нет в наличии",
            "відсутній",
            "відсутня",
        ]

        for word in unavailable_words:

            if word in text:
                return "Немає в наявності"

        return "В наявності"

    # --------------------------------------------------------
    # Schema availability
    # --------------------------------------------------------

    availability = soup.select_one(
        "link[itemprop='availability']"
    )

    if availability:

        value = (
            availability.get("href")
            or availability.get("content")
            or ""
        )

        value = str(value).casefold()

        if "outofstock" in value:
            return "Немає в наявності"

        if "instock" in value:
            return "В наявності"

    return "В наявності"


# ============================================================
# COLORS
# ============================================================

def get_product_colors(soup):

    """
    Каждый цвет — отдельная строка.

    Например:

    SKU | COLOR
    697 | Блакитний
    697 | Зелений
    697 | Рожевий
    697 | Чорний
    697 | Фіолетовий
    """

    if not soup:
        return []

    colors = []

    fieldsets = soup.select(
        "#attributes .attribute_fieldset"
    )

    for fieldset in fieldsets:

        legend = fieldset.select_one(
            "legend"
        )

        legend_text = ""

        if legend:
            legend_text = clean(
                legend.get_text(
                    " ",
                    strip=True,
                )
            ).casefold()

        if "колір" not in legend_text:
            continue

        inputs = fieldset.select(
            "input.attribute_radio"
        )

        for input_element in inputs:

            color_id = input_element.get("value")

            if not color_id:
                continue

            color_id = str(color_id).strip()

            color_name = ""

            # ------------------------------------------------
            # Пытаемся найти текст рядом с input
            # ------------------------------------------------

            parent = input_element.parent

            if parent:

                text = clean(
                    parent.get_text(
                        " ",
                        strip=True,
                    )
                )

                text = re.sub(
                    r"\bchecked\b",
                    "",
                    text,
                    flags=re.I,
                )

                text = clean(text)

                if text:
                    color_name = text

            # ------------------------------------------------
            # Если не нашли — ищем li
            # ------------------------------------------------

            if not color_name:

                li = input_element.find_parent("li")

                if li:

                    text = clean(
                        li.get_text(
                            " ",
                            strip=True,
                        )
                    )

                    text = re.sub(
                        r"\bchecked\b",
                        "",
                        text,
                        flags=re.I,
                    )

                    color_name = clean(text)

            if not color_name:
                continue

            colors.append({
                "id": color_id,
                "name": color_name,
            })

    # Убираем дубликаты
    result = []
    seen = set()

    for color in colors:

        key = (
            str(color["id"]),
            normalize_sku(color["name"]),
        )

        if key in seen:
            continue

        seen.add(key)
        result.append(color)

    return result


# ============================================================
# PRODUCT
# ============================================================

def parse_product(
    product_url,
    price_map,
):

    clean_url = base_product_url(product_url)

    soup = get_soup(clean_url)

    if not soup:
        return []

    title = get_product_title(soup)
    sku = get_product_sku(soup)
    status = get_product_status(soup)

    normalized_sku = normalize_sku(sku)

    price = None

    if normalized_sku:

        price = price_map.get(
            normalized_sku
        )

    if price is None:
        price_value = ""
    else:
        price_value = price

    colors = get_product_colors(soup)

    rows = []

    # ========================================================
    # БЕЗ ЦВЕТОВ
    # ========================================================

    if not colors:

        rows.append({
            "SKU": sku,
            "TITLE": title,
            "PRICE USD": price_value,
            "STATUS": status,
            "COLOR": "",
            "URL": clean_url,
        })

        return rows

    # ========================================================
    # С ЦВЕТАМИ
    # ========================================================

    for color in colors:

        color_id = color["id"]
        color_name = color["name"]

        slug = translit_slug(color_name)

        color_url = (
            f"{clean_url}"
            f"#/{color_id}-kolir-{slug}"
        )

        rows.append({
            "SKU": sku,
            "TITLE": title,
            "PRICE USD": price_value,
            "STATUS": status,
            "COLOR": color_name,
            "URL": color_url,
        })

    return rows


# ============================================================
# EXCEL
# ============================================================

def save_excel(rows):

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "ARIZONE Sports"

    headers = [
        "SKU",
        "TITLE",
        "PRICE USD",
        "STATUS",
        "COLOR",
        "URL",
    ]

    worksheet.append(headers)

    for row in rows:

        worksheet.append([
            row.get("SKU", ""),
            row.get("TITLE", ""),
            row.get("PRICE USD", ""),
            row.get("STATUS", ""),
            row.get("COLOR", ""),
            row.get("URL", ""),
        ])

    # --------------------------------------------------------
    # Форматирование
    # --------------------------------------------------------

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 60
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 22
    worksheet.column_dimensions["E"].width = 20
    worksheet.column_dimensions["F"].width = 90

    for cell in worksheet[1]:

        cell.font = cell.font.copy(
            bold=True
        )

    for row in worksheet.iter_rows(
        min_row=2,
        min_col=3,
        max_col=3,
    ):

        cell = row[0]

        if isinstance(cell.value, (int, float)):
            cell.number_format = '0.##'

    workbook.save(FILE_PATH)

    print("")
    print(f"💾 Excel сохранён:")
    print(f"   {FILE_PATH}")


# ============================================================
# MAIN
# ============================================================

def main():

    print("")
    print("🔥 BOT STARTED")
    print("🚀 BOT RUNNING")
    print("🚀 STARTED Самокаты ARIZONE Sports")
    print("📄 SCRIPT: parsers/ARIZONE_Sports/run.py")
    print("")

    if not create_lock():
        return

    save_status(
        "running",
        0,
        0,
        "Загрузка цен из Google Sheets",
    )

    try:

        # ====================================================
        # GOOGLE SHEETS
        # ====================================================

        price_map = get_google_sheet_price_map()

        if not price_map:

            print("")
            print("=" * 70)
            print("❌ ПАРСЕР ОСТАНОВЛЕН")
            print("=" * 70)
            print(
                "Google Sheets не дала ни одной цены."
            )
            print(
                "Не создаём Excel с пустым PRICE USD."
            )
            print("=" * 70)

            save_status(
                "error",
                0,
                0,
                "Не удалось загрузить цены из Google Sheets",
            )

            return

        # ====================================================
        # CATEGORIES
        # ====================================================

        categories = get_categories()

        if not categories:

            print("❌ Категории не найдены.")

            save_status(
                "error",
                0,
                0,
                "Категории не найдены",
            )

            return

        if CATEGORY_LIMIT:
            categories = categories[:CATEGORY_LIMIT]

        print("")
        print(
            f"📂 Будет обработано категорий: "
            f"{len(categories)}"
        )

        # ====================================================
        # ALL PRODUCTS
        # ====================================================

        all_product_urls = []
        seen_products = set()

        for category_index, category in enumerate(
            categories,
            start=1,
        ):

            save_status(
                "running",
                category_index,
                len(categories),
                (
                    f"Категория: "
                    f"{category['name']}"
                ),
            )

            print("")
            print("=" * 70)
            print(
                f"📂 КАТЕГОРИЯ "
                f"{category_index}/{len(categories)}"
            )
            print(
                f"   {category['name']}"
            )
            print("=" * 70)

            product_urls = get_category_products(
                category["url"]
            )

            for product_url in product_urls:

                base_url = base_product_url(
                    product_url
                )

                if base_url in seen_products:
                    continue

                seen_products.add(base_url)
                all_product_urls.append(base_url)

        print("")
        print("=" * 70)
        print(
            f"🛒 УНИКАЛЬНЫХ ТОВАРОВ: "
            f"{len(all_product_urls)}"
        )
        print("=" * 70)

        # ====================================================
        # PRODUCT LIMIT
        # ====================================================

        products_to_parse = all_product_urls

        if PRODUCT_LIMIT:
            products_to_parse = all_product_urls[
                :PRODUCT_LIMIT
            ]

        print(
            f"🔎 К обработке товаров: "
            f"{len(products_to_parse)}"
        )

        # ====================================================
        # PARSE PRODUCTS
        # ====================================================

        final_rows = []

        total_products = len(products_to_parse)

        for index, product_url in enumerate(
            products_to_parse,
            start=1,
        ):

            print("")
            print(
                f"🔎 ТОВАР "
                f"{index}/{total_products}"
            )
            print(
                f"   {product_url}"
            )

            save_status(
                "running",
                index,
                total_products,
                f"Товар {index}/{total_products}",
            )

            try:

                rows = parse_product(
                    product_url,
                    price_map,
                )

                for row in rows:

                    final_rows.append(row)

                    sku = row.get(
                        "SKU",
                        "",
                    )

                    price = row.get(
                        "PRICE USD",
                        "",
                    )

                    color = row.get(
                        "COLOR",
                        "",
                    )

                    print(
                        f"   ✓ "
                        f"{sku} | "
                        f"{color} | "
                        f"${price}"
                    )

            except Exception as e:

                print(
                    f"   ❌ PRODUCT ERROR: "
                    f"{e}"
                )

        # ====================================================
        # DEDUPE ROWS
        # ====================================================

        unique_rows = []
        seen_rows = set()

        for row in final_rows:

            sku = normalize_sku(
                row.get("SKU", "")
            )

            color = normalize_sku(
                row.get("COLOR", "")
            )

            url = base_product_url(
                row.get("URL", "")
            )

            if sku:

                key = (
                    sku,
                    color,
                )

            else:

                key = (
                    url,
                    color,
                )

            if key in seen_rows:
                continue

            seen_rows.add(key)
            unique_rows.append(row)

        final_rows = unique_rows

        # ====================================================
        # PRICE DIAGNOSTICS
        # ====================================================

        priced_count = sum(
            1
            for row in final_rows
            if row.get("PRICE USD") != ""
        )

        unpriced_count = len(final_rows) - priced_count

        print("")
        print("=" * 70)
        print("💰 ПРОВЕРКА ЦЕН")
        print("=" * 70)
        print(
            f"💵 Строк с ценой: "
            f"{priced_count}"
        )
        print(
            f"⚠️ Строк без цены: "
            f"{unpriced_count}"
        )

        if unpriced_count:

            print("")
            print(
                "🔎 Артикулы без найденной цены:"
            )

            shown = 0
            shown_skus = set()

            for row in final_rows:

                if row.get("PRICE USD") != "":
                    continue

                sku = row.get("SKU", "")

                normalized = normalize_sku(
                    sku
                )

                if normalized in shown_skus:
                    continue

                shown_skus.add(normalized)

                print(
                    f"   ❌ {sku}"
                )

                shown += 1

                if shown >= 30:
                    break

        print("=" * 70)

        # ====================================================
        # SAVE EXCEL
        # ====================================================

        save_excel(final_rows)

        # ====================================================
        # FINISH
        # ====================================================

        save_status(
            "done",
            len(final_rows),
            len(final_rows),
            (
                f"Готово. "
                f"Строк: {len(final_rows)}, "
                f"с ценой: {priced_count}"
            ),
        )

        print("")
        print("=" * 70)
        print("✅ PARSER FINISHED")
        print("=" * 70)
        print(
            f"📦 Строк в Excel: "
            f"{len(final_rows)}"
        )
        print(
            f"💰 С ценой USD: "
            f"{priced_count}"
        )
        print(
            f"⚠️ Без цены: "
            f"{unpriced_count}"
        )
        print("=" * 70)

    except Exception as e:

        print("")
        print("=" * 70)
        print("❌ FATAL ERROR")
        print("=" * 70)
        print(e)
        print("=" * 70)

        save_status(
            "error",
            0,
            0,
            str(e),
        )

    finally:

        remove_lock()


# ============================================================
# START
# ============================================================

if __name__ == "__main__":
    main()
