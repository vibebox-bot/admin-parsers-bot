import os
import re
import json
import time
import random
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook, load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from datetime import datetime

print("🟢 FILE LOADED: parser script imported")

# =====================================================
# НАСТРОЙКИ
# =====================================================

BASE_URL = "https://lambix.prom.ua"
CATALOG_URL = "https://lambix.prom.ua/ua/product_list"

OUTPUT_DIR = os.path.abspath("output/220_Kithen_Plus")
FILE_PATH = os.path.join(OUTPUT_DIR, "220_Kithen_Plus_LIVE.xlsx")
STATUS_PATH = os.path.join(OUTPUT_DIR, "status.json")
LOCK_FILE = os.path.join(OUTPUT_DIR, "lock.txt")

# CATEGORY_LIMIT = None
CATEGORY_LIMIT = 1

import sys
USER = sys.argv[1] if len(sys.argv) > 1 else ""

os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/137.0 Safari/537.36"
    )
}

# =====================================================
# SESSION
# =====================================================

retry = Retry(
    total=5,
    connect=5,
    read=5,
    backoff_factor=1,
    status_forcelist=[429, 500, 502, 503, 504],
)

adapter = HTTPAdapter(max_retries=retry)

session = requests.Session()
session.headers.update(HEADERS)
session.mount("https://", adapter)
session.mount("http://", adapter)

# =====================================================
# STATUS
# =====================================================

def load_status():
    if not os.path.exists(STATUS_PATH):
        return {"category": 0}
    with open(STATUS_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_status(category_index, running=False, progress=0, canceled=False, success=False):
    data = {
        "category": category_index,
        "running": running,
        "progress": progress,
        "canceled": canceled,
        "success": success,
        "user": USER,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    with open(STATUS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

# =====================================================
# EXCEL
# =====================================================

def get_workbook():
    if os.path.exists(FILE_PATH):
        wb = load_workbook(FILE_PATH)
        ws = wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.append(["Название", "Артикул", "Цена", "Наличие", "Ссылка"])
    wb.save(FILE_PATH)
    return wb, ws

# =====================================================
# REQUEST
# =====================================================

def get_soup(url):
    print(f"🌍 REQUEST: {url}")
    attempt = 0

    while attempt < 5:
        attempt += 1
        try:
            r = session.get(url, timeout=15)
            print(f"🌐 {url} -> {r.status_code}")

            if r.status_code == 200:
                return BeautifulSoup(r.text, "html.parser")

        except Exception as e:
            print(e)

        time.sleep(random.uniform(2, 4))

    return None

# =====================================================
# КАТЕГОРИИ
# =====================================================

def get_categories():
    soup = get_soup(CATALOG_URL)
    if soup is None:
        print("❌ Не удалось загрузить каталог")
        return []

    categories = []
    blocks = soup.select("ul.b-product-groups-gallery li")

    for block in blocks:
        a = block.select_one("a.b-product-groups-gallery__title")
        if not a:
            continue

        href = a.get("href")
        if not href:
            continue

        categories.append(urljoin(BASE_URL, href))

    if CATEGORY_LIMIT:
        categories = categories[:CATEGORY_LIMIT]

    print(f"📂 Категорий найдено: {len(categories)}")
    for i, c in enumerate(categories):
        print(f"{i+1}. {c}")

    return categories


def get_subcategories(category_url):
    soup = get_soup(category_url)
    subs = []

    blocks = soup.select("ul.b-product-groups-gallery li")

    for block in blocks:
        a = block.select_one("a.b-product-groups-gallery__title")
        if not a:
            continue

        href = a.get("href")
        if href:
            subs.append(urljoin(BASE_URL, href))

    return subs


def process_category(category_url, ws, wb):

    if category_url in visited_categories:
        return

    visited_categories.add(category_url)

    print(f"📂 CATEGORY: {category_url}")

    subcategories = get_subcategories(category_url)

    print(f"Подкатегорий найдено: {len(subcategories)}")

    for s in subcategories:
        print("  ↳", s)

    if subcategories:

        print(f"📁 Найдено подкатегорий: {len(subcategories)}")

        for sub in subcategories:
            process_category(sub, ws, wb)

        return

    pages = get_pages(category_url)

    for page in pages:

        print("\nСтраница:")
        print(page)

        products = get_products(page)

        for product_url in products:

            try:

                product = parse_product(product_url)

                save_product(ws, wb, product)

                time.sleep(random.uniform(0.5, 1.2))

            except Exception as e:

                print("Ошибка товара:")
                print(product_url)
                print(e)

        time.sleep(random.uniform(1, 2))

# =====================================================
# ПАГИНАТОР
# =====================================================

def get_pages(category_url):
    soup = get_soup(category_url)
    last_page = 1

    pager = soup.select(".b-pager a")

    for a in pager:
        txt = a.get_text(strip=True)
        if txt.isdigit():
            last_page = max(last_page, int(txt))

    pages = []

    for i in range(1, last_page + 1):
        if i == 1:
            pages.append(category_url)
        else:
            pages.append(category_url.rstrip("/") + f"/page_{i}")

    print(f"Страниц: {len(pages)}")
    return pages

# =====================================================
# ТОВАРЫ
# =====================================================

def get_products(page_url):
    soup = get_soup(page_url)
    products = []

    cards = soup.select("li[data-product-id]")

    for card in cards:
        a = card.select_one("a.b-product-gallery__title")
        if not a:
            continue

        href = a.get("href")
        if href:
            products.append(urljoin(BASE_URL, href))

    print(f"🛒 Товаров на странице: {len(products)}")

    if len(products) == 0:
        print("⚠️ НЕТ ТОВАРОВ — возможно сайт блокирует или сломался селектор")

    return products

# =====================================================
# ПАРС ТОВАРА
# =====================================================

def clean_price(text):
    if not text:
        return ""
    text = text.replace("\xa0", "")
    text = text.replace("₴", "")
    text = text.replace("/шт.", "")
    return text.strip()


def get_text(el):
    if not el:
        return ""
    return el.get_text(" ", strip=True)


def parse_product(url):
    soup = get_soup(url)

    title = get_text(soup.select_one('span[data-qaid="product_name"]'))
    sku = get_text(soup.select_one('span[data-qaid="product_code"]'))

    price = ""
    price_tag = soup.select_one('span[data-qaid="wholesale_price"]')
    if price_tag:
        price = clean_price(price_tag.get_text(strip=True))

    availability = ""
    state = soup.select_one('[data-qaid="presence_data"]')
    if state:
        availability = state.get_text(" ", strip=True)

    return {
        "title": title,
        "sku": sku,
        "price": price,
        "availability": availability,
        "url": url
    }

# =====================================================
# EXCEL SAVE
# =====================================================

def save_product(ws, wb, product):
    ws.append([
        product["title"],
        product["sku"],
        product["price"],
        product["availability"],
        product["url"]
    ])
    wb.save(FILE_PATH)
    print(product["title"])

# =====================================================
# MAIN
# =====================================================

def main():
    print("🔥 ENTER MAIN()")
    print("🚀 PARSER STARTED")
    print(f"👤 USER: {USER}")

    with open(LOCK_FILE, "w", encoding="utf-8") as f:
        f.write(str(os.getpid()))

    save_status(0, running=True)

    status = load_status()
    wb, ws = get_workbook()

    categories = get_categories()
    start_category = status.get("category", 0)

    for category_index in range(start_category, len(categories)):
        category_url = categories[category_index]

        print("\n" + "=" * 70)
        print(f"Категория {category_index + 1}/{len(categories)}")
        print(category_url)

        process_category(category_url, ws, wb)

    print("\n==========================")
    print("ПАРСИНГ ЗАВЕРШЕН")
    print("==========================")

# =====================================================
# START
# =====================================================

def run_parser():
    main()


if __name__ == "__main__":
    run_parser()
