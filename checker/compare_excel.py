import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font

# =========================
# ФАЙЛЫ
# =========================

CRM_FILE = "checker/export-2026-07-12_00-14-43.xlsx"
MELAD_FILE = "output/Melad/Melad_LIVE.xlsx"
RESULT_FILE = "output/Melad/Melad_CHECK.xlsx"



YELLOW = PatternFill(fill_type="solid", fgColor="FFF59D")
RED = PatternFill(fill_type="solid", fgColor="F8BBD0")
ORANGE = PatternFill(fill_type="solid", fgColor="FFD180")
GREEN = PatternFill(fill_type="solid", fgColor="C8E6C9")

BOLD = Font(bold=True)

# =========================
# CRM EXCEL
# =========================

def get_salesdrive():

    wb = load_workbook(CRM_FILE, data_only=True)
    ws = wb.active

    headers = {}

    for i, cell in enumerate(ws[1], start=1):
        headers[str(cell.value).strip()] = i

    print("📥 SalesDrive Excel")
    print(headers)

    supplier_col = headers["Постачальник"]
    barcode_col = headers["Штрихкод"]
    note_col = headers["Нотатка"]
    cost_col = headers["Собівартість"]
    price_col = headers["Ціна"]

    products = {}

    for row in ws.iter_rows(min_row=2, values_only=True):

        supplier = row[supplier_col - 1]

        if str(supplier).strip() != "Melad":
            continue

        article = ""

        barcode = row[barcode_col - 1]
        note = row[note_col - 1]

        if barcode not in ("", None):
            article = str(barcode).strip()

        elif note not in ("", None):
            article = str(note).strip()

        if article == "":
            continue

        cost = row[cost_col - 1]

        try:
            site_price = float(row[price_col - 1])
        except:
            site_price = 0

        if site_price <= 1:
            stock = 0
        else:
            stock = 1

        products[article] = {
            "cost": cost,
            "stock": stock,
            "site_price": site_price
        }

    print("Melad в CRM:", len(products))

    return products


# =========================
# MELAD PARSER
# =========================

def get_melad():

    wb = load_workbook(MELAD_FILE, data_only=True)
    ws = wb.active

    products = {}

    for row in ws.iter_rows(min_row=2, values_only=True):

        name = row[0]
        price = row[1]
        article = row[2]
        stock = row[3]
        url = row[4]

        if not article:
            continue

        article = str(article).strip()

        # Цена
        price = str(price).replace("$", "")
        price = price.replace("грн", "")
        price = price.replace(",", ".")
        price = price.strip()

        try:
            price = float(price)
        except:
            price = 0

       # Наличие
        stock = str(stock).strip().lower()
        
        if "в корзину" in stock:
            stock = 1
        else:
            stock = 0

        products[article] = {
            "name": name,
            "price": price,
            "stock": stock,
            "url": url
        }

    print("📥 Melad parser")
    print("Melad товаров:", len(products))

    return products

# =========================
# MAIN
# =========================

def main():

    crm = get_salesdrive()
    melad = get_melad()

    wb = Workbook()
    ws = wb.active
    ws.title = "Melad"


    ws.append([
        "Действие",
        "Артикул",
        "Название",
        "Было (CRM)",
        "Стало (Melad)",
        "URL"
    ])

    for cell in ws[1]:
        cell.font = BOLD


    changed_price = 0
    changed_stock = 0
    ok = 0
    missing = 0

    for article, item in melad.items():

        crm_item = crm.get(article)

        if crm_item is None:

            ws.append([
                "🆕 Создать товар",
                article,
                item["name"],
                "-",
                f"{item['price']}$ / {'Есть' if item['stock'] else 'Нет'}",
                item["url"]
            ])
            
            row = ws.max_row
            
            for cell in ws[row]:
                cell.fill = GREEN
        

            missing += 1
            continue

        # ---------- Цена ----------

        parser_price = item["price"]

        try:
            crm_cost = float(str(crm_item["cost"]).replace(",", "."))
        except:
            crm_cost = 0

        # ---------- Наличие ----------

        parser_stock = 0 if str(item["stock"]).strip() == "0" else 1
        crm_stock = crm_item["stock"]

        price_changed = abs(parser_price - crm_cost) > 0.001
        stock_changed = parser_stock != crm_stock
        
        if not price_changed and not stock_changed:
            ok += 1
            continue
        
        if price_changed and stock_changed:
            action = "💲📦 Обновить цену и наличие"
            changed_price += 1
            changed_stock += 1
        
        elif price_changed:
            action = "💲 Обновить себестоимость"
            changed_price += 1
        
        else:
            action = "📦 Обновить наличие"
            changed_stock += 1
        
        crm_stock_text = "Есть" if crm_stock else "Нет"
        melad_stock_text = "Есть" if parser_stock else "Нет"
        
        if price_changed and stock_changed:
        
            old_value = f"{crm_cost}$ / {crm_stock_text}"
            new_value = f"{parser_price}$ / {melad_stock_text}"
        
        elif price_changed:
        
            old_value = f"{crm_cost}$"
            new_value = f"{parser_price}$"
        
        else:
        
            old_value = crm_stock_text
            new_value = melad_stock_text

       
        ws.append([
            action,
            article,
            item["name"],
            old_value,
            new_value,
            item["url"]
        ])

        row = ws.max_row
    
        fill = None
        
        if action == "💲 Обновить себестоимость":
            fill = YELLOW
        
        elif action == "📦 Обновить наличие":
            fill = RED
        
        elif action == "💲📦 Обновить цену и наличие":
            fill = ORANGE
        
        elif action == "🆕 Создать товар":
            fill = GREEN
        
        if fill:
            for cell in ws[row]:
                cell.fill = fill
    

    wb.save(RESULT_FILE)

    print()
    print("==========")
    print("OK:", ok)
    print("Цена:", changed_price)
    print("Наличие:", changed_stock)
    print("Нет в CRM:", missing)
    print("==========")
    print(RESULT_FILE)

if __name__ == "__main__":
    main()
