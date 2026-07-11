import os
from openpyxl import load_workbook, Workbook


SALES_FILE = "checker/export-2026-07-12_00-14-43.xlsx"

MELAD_FILE = "output/Melad/Melad_LIVE.xlsx"

RESULT_FILE = "output/Melad/Melad_CHECK.xlsx"


# =========================
# SALES DRIVE EXCEL
# =========================

def get_salesdrive():

    wb = load_workbook(SALES_FILE, data_only=True)
    ws = wb.active

    headers = {}

    for col in range(1, ws.max_column + 1):
        name = ws.cell(1, col).value

        if name:
            headers[name] = col


    print("Колонки найдены:")
    print(headers)


    products = {}


    for row in range(2, ws.max_row + 1):

        supplier = ws.cell(
            row,
            headers["Постачальник"]
        ).value


        if supplier != "Melad":
            continue


        sku = ws.cell(
            row,
            headers["SKU"]
        ).value


        if not sku:
            continue


        products[str(sku)] = {

            "name":
                ws.cell(row, headers["Товар/Послуга"]).value,

            "cost":
                ws.cell(row, headers["Собівартість"]).value,

            "cost_currency":
                ws.cell(row, headers["Собівартість - Валюта"]).value,

            "stock":
                ws.cell(row, headers["Залишок на складі"]).value,

            "price":
                ws.cell(row, headers["Ціна"]).value,

            "note":
                ws.cell(row, headers["Нотатка"]).value
        }


    return products



# =========================
# MELAD
# =========================

def get_melad():

    wb = load_workbook(MELAD_FILE, data_only=True)
    ws = wb.active


    products = {}


    for row in ws.iter_rows(min_row=2, values_only=True):

        name, price, article, stock, url = row


        if article:

            products[str(article)] = {

                "name": name,
                "price": price,
                "stock": stock,
                "url": url

            }


    return products



# =========================
# COMPARE
# =========================

def main():

    print("📥 SalesDrive Excel")

    crm = get_salesdrive()

    print(
        "Melad в CRM:",
        len(crm)
    )


    print("📥 Melad parser")

    melad = get_melad()

    print(
        "Melad товаров:",
        len(melad)
    )


    wb = Workbook()

    ws = wb.active

    ws.title = "Проверка"


    ws.append([

        "Товар Melad",
        "SKU",
        "Поставщик",
        "Себестоимость",
        "Остаток CRM",
        "Цена CRM",
        "Цена сайта",
        "Статус",
        "URL"

    ])


    count = 0


    for sku, item in melad.items():

        crm_item = crm.get(str(sku))


        if crm_item:


            stock = crm_item["stock"]


            if stock in [0, "0", None]:
                status = "❌ НЕТ В НАЛИЧИИ"
            else:
                status = "✅ ЕСТЬ"


            ws.append([

                item["name"],

                sku,

                "Melad",

                crm_item["cost"],

                stock,

                crm_item["price"],

                item["price"],

                status,

                item["url"]

            ])


            count += 1



    os.makedirs(
        "output/Melad",
        exist_ok=True
    )


    wb.save(
        RESULT_FILE
    )


    print()
    print("ГОТОВО:", count)
    print(RESULT_FILE)



if __name__ == "__main__":
    main()
