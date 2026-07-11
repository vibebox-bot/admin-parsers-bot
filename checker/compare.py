import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font


YML_URL = "https://kindlytech.salesdrive.me/export/yml/export.yml?publicKey=YEWxvIKV_z6Hjx4-zqWiGLmmsFAS05TLQQ23qZbeoR_2UjOCNEtx-QxFfP0JFfUv45Q"


MELAD_FILE = "output/Melad/Melad_LIVE.xlsx"

RESULT_FILE = "output/Melad/Melad_CHECK.xlsx"


# =========================
# SALES DRIVE
# =========================

def get_salesdrive():

    r = requests.get(
        YML_URL,
        timeout=60
    )

    r.raise_for_status()

    root = ET.fromstring(r.text)

    products = {}

    for offer in root.findall(".//offer"):

        note = offer.findtext("note", "")
        barcode = offer.findtext("barcode", "")

        note = note.strip() if note else ""
        barcode = barcode.strip() if barcode else ""

        keys = []

        if note:
            keys.append(note)

        if barcode:
            keys.append(barcode)


        for key in keys:

            products[key] = {

                "name": offer.findtext("name", ""),

                "stock": offer.findtext(
                    "quantity_in_stock",
                    "0"
                ),

                "price": offer.findtext(
                    "price",
                    ""
                )
            }


    return products



# =========================
# MELAD
# =========================

def get_melad():

    wb = load_workbook(
        MELAD_FILE
    )

    ws = wb.active

    products = {}

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        name, price, article, stock, url = row


        if article:

            products[str(article).strip()] = {

                "name": name,
                "price": price,
                "url": url
            }


    return products



# =========================
# COMPARE
# =========================

def main():

    print("📥 SalesDrive...")
    crm = get_salesdrive()

    print(
        "CRM товаров:",
        len(crm)
    )


    print("📥 Melad...")
    melad = get_melad()

    print(
        "Melad товаров:",
        len(melad)
    )


    wb = Workbook()


    # =====================
    # НАЙДЕННЫЕ
    # =====================

    ws = wb.active
    ws.title = "Проверка"


    ws.append([
        "Название",
        "Арт поставщика",
        "Цена Melad",
        "Цена CRM",
        "Наличие CRM",
        "Статус"
    ])


    # =====================
    # НЕ НАЙДЕНЫ
    # =====================

    ws2 = wb.create_sheet(
        "Не найдено"
    )

    ws2.append([
        "Название",
        "Арт поставщика",
        "URL"
    ])


    found = 0
    not_found = 0


    for article, item in melad.items():


        crm_item = crm.get(article)


        if crm_item:

            crm_price = str(
                crm_item["price"]
            )


            crm_stock = str(
                crm_item["stock"]
            )


            if crm_stock == "0" or crm_price == "1":

                status = "❌ НЕТ В НАЛИЧИИ"

            else:

                status = "✅ ЕСТЬ"


            ws.append([

                item["name"],

                article,

                item["price"],

                crm_price,

                crm_stock,

                status
            ])


            found += 1


        else:


            ws2.append([

                item["name"],

                article,

                item["url"]

            ])

            not_found += 1



    # ширина колонок

    for sheet in [ws, ws2]:

        for col in sheet.columns:

            sheet.column_dimensions[
                col[0].column_letter
            ].width = 25


    # жирный заголовок

    for sheet in [ws, ws2]:

        for cell in sheet[1]:

            cell.font = Font(
                bold=True
            )


    wb.save(
        RESULT_FILE
    )


    print("----------------")
    print("НАЙДЕНО:", found)
    print("НЕ НАЙДЕНО:", not_found)
    print(RESULT_FILE)



if __name__ == "__main__":
    main()
